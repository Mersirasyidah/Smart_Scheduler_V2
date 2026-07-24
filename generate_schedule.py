import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ==========================================
# 1. INISIALISASI WORKBOOK
# ==========================================
wb = openpyxl.Workbook()
wb.remove(wb.active)  # Hapus sheet bawaan (Sheet default)

# ==========================================
# 2. PEMETAAN GURU & MAPEL PER KELAS
# ==========================================
teaching_assignment = {
    '7A': {
        'IPA': 'Purwanto, S.Pd.',
        'B.Indonesia': 'Asti Am Rini, S.Pd.',
        'Matematika': 'Eny Widiyanti, S.Pd.',
        'B.Inggris': 'Wesda Ayu Rahmadani',
        'PJOK': 'Cholid Dalyanto, S.Pd.Kor.',
        'PAI': 'Anik Mulyani, S.Ag.',
        'P.Pancasila': 'Christina Dwi Ayu W.',
        'IPS': 'Rizal Rahmanto, S.Pd.',
        'Informatika': 'Anggiyani Fabilah P.',
        'Prakarya': 'Thiara Maharani, S.Pd.',
        'B.Jawa': 'Fitri Lestari, S.S.',
    },
    '7B': {
        'IPA': 'Purwanto, S.Pd.',
        'B.Indonesia': 'Asti Am Rini, S.Pd.',
        'Matematika': 'Eny Widiyanti, S.Pd.',
        'B.Inggris': 'Wesda Ayu Rahmadani',
        'PJOK': 'Cholid Dalyanto, S.Pd.Kor.',
        'PAI': 'Anik Mulyani, S.Ag.',
        'P.Pancasila': 'Christina Dwi Ayu W.',
        'IPS': 'Rizal Rahmanto, S.Pd.',
        'Informatika': 'Anggiyani Fabilah P.',
        'Prakarya': 'Thiara Maharani, S.Pd.',
        'B.Jawa': 'Fitri Lestari, S.S.',
    },
    '7C': {
        'IPA': 'Purwanto, S.Pd.',
        'B.Indonesia': 'Asti Am Rini, S.Pd.',
        'Matematika': 'Eny Widiyanti, S.Pd.',
        'B.Inggris': 'Wesda Ayu Rahmadani',
        'PJOK': 'Cholid Dalyanto, S.Pd.Kor.',
        'PAI': 'Anik Mulyani, S.Ag.',
        'P.Pancasila': 'Christina Dwi Ayu W.',
        'IPS': 'Rizal Rahmanto, S.Pd.',
        'Informatika': 'Anggiyani Fabilah P.',
        'Prakarya': 'Thiara Maharani, S.Pd.',
        'B.Jawa': 'Fitri Lestari, S.S.',
    },
    '7D': {
        'IPA': 'Purwanto, S.Pd.',
        'B.Indonesia': 'Asti Am Rini, S.Pd.',
        'Matematika': 'Eny Widiyanti, S.Pd.',
        'B.Inggris': 'Wesda Ayu Rahmadani',
        'PJOK': 'Luthfan Qaedi W.',
        'PAI': 'Anik Mulyani, S.Ag.',
        'P.Pancasila': 'Feni Dwimartanti, S.Pd.',
        'IPS': 'Rizal Rahmanto, S.Pd.',
        'Informatika': 'Anggiyani Fabilah P.',
        'Prakarya': 'Thiara Maharani, S.Pd.',
        'B.Jawa': 'Fitri Lestari, S.S.',
    },
    '7E': {
        'IPA': 'Purwanto, S.Pd.',
        'B.Indonesia': 'Asti Am Rini, S.Pd.',
        'Matematika': 'Eny Widiyanti, S.Pd.',
        'B.Inggris': 'Wesda Ayu Rahmadani',
        'PJOK': 'Luthfan Qaedi W.',
        'PAI': 'Anik Mulyani, S.Ag.',
        'P.Pancasila': 'Feni Dwimartanti, S.Pd.',
        'IPS': 'Rizal Rahmanto, S.Pd.',
        'Informatika': 'Anggiyani Fabilah P.',
        'Prakarya': 'Thiara Maharani, S.Pd.',
        'B.Jawa': 'Fitri Lestari, S.S.',
    },
    '8A': {
        'IPA': 'Anisa Safera Proborini',
        'B.Indonesia': 'Milenia, S.Pd.',
        'Matematika': 'Ami Royati, S.Pd.',
        'B.Inggris': 'Rizka Diestriana, S.Pd.',
        'PJOK': 'Cholid Dalyanto, S.Pd.Kor.',
        'PAI': 'Anggi Supriyadi, S.Hum',
        'P.Pancasila': 'Christina Dwi Ayu W.',
        'IPS': 'Rizal Rahmanto, S.Pd.',
        'Informatika': 'Mersi, S.T.',
        'Seni Budaya': 'Yoma Septiantika, S.Pd.',
        'B.Jawa': 'Fitri Lestari, S.S.',
    },
    '8B': {
        'IPA': 'Anisa Safera Proborini',
        'B.Indonesia': 'Milenia, S.Pd.',
        'Matematika': 'Ami Royati, S.Pd.',
        'B.Inggris': 'Rizka Diestriana, S.Pd.',
        'PJOK': 'Cholid Dalyanto, S.Pd.Kor.',
        'PAI': 'Anggi Supriyadi, S.Hum',
        'P.Pancasila': 'Christina Dwi Ayu W.',
        'IPS': 'Rizal Rahmanto, S.Pd.',
        'Informatika': 'Mersi, S.T.',
        'Seni Budaya': 'Yoma Septiantika, S.Pd.',
        'B.Jawa': 'Fitri Lestari, S.S.',
    },
    '8C': {
        'IPA': 'Anisa Safera Proborini',
        'B.Indonesia': 'Milenia, S.Pd.',
        'Matematika': 'Ami Royati, S.Pd.',
        'B.Inggris': 'Rizka Diestriana, S.Pd.',
        'PJOK': 'Cholid Dalyanto, S.Pd.Kor.',
        'PAI': 'Anik Mulyani, S.Ag.',
        'P.Pancasila': 'Christina Dwi Ayu W.',
        'IPS': 'Budi Prasetya, S.Pd.',
        'Informatika': 'Anggiyani Fabilah P.',
        'Seni Budaya': 'Yoma Septiantika, S.Pd.',
        'B.Jawa': 'Thiara Maharani, S.Pd.',
    },
    '8D': {
        'IPA': 'Anisa Safera Proborini',
        'B.Indonesia': 'Milenia, S.Pd.',
        'Matematika': 'Ami Royati, S.Pd.',
        'B.Inggris': 'Rizka Diestriana, S.Pd.',
        'PJOK': 'Cholid Dalyanto, S.Pd.Kor.',
        'PAI': 'Anik Mulyani, S.Ag.',
        'P.Pancasila': 'Christina Dwi Ayu W.',
        'IPS': 'Budi Prasetya, S.Pd.',
        'Informatika': 'Anggiyani Fabilah P.',
        'Seni Budaya': 'Yoma Septiantika, S.Pd.',
        'B.Jawa': 'Thiara Maharani, S.Pd.',
    },
    '8E': {
        'IPA': 'Anisa Safera Proborini',
        'B.Indonesia': 'Milenia, S.Pd.',
        'Matematika': 'Ami Royati, S.Pd.',
        'B.Inggris': 'Rizka Diestriana, S.Pd.',
        'PJOK': 'Cholid Dalyanto, S.Pd.Kor.',
        'PAI': 'Anik Mulyani, S.Ag.',
        'P.Pancasila': 'Christina Dwi Ayu W.',
        'IPS': 'Budi Prasetya, S.Pd.',
        'Informatika': 'Anggiyani Fabilah P.',
        'Seni Budaya': 'Yoma Septiantika, S.Pd.',
        'B.Jawa': 'Thiara Maharani, S.Pd.',
    },
    '9A': {
        'IPA': 'Umi Kulstum, S.Pd.',
        'B.Indonesia': 'Maftuhah Rahayu, S.Pd.',
        'Matematika': 'Sri Purwanti, S.Pd.',
        'B.Inggris': 'Heni Purwaningsih, S.Pd.',
        'PJOK': 'Luthfan Qaedi W.',
        'PAI': 'Nurkhasanah, S.Ag.',
        'P.Pancasila': 'Feni Dwimartanti, S.Pd.',
        'IPS': 'Budi Prasetya, S.Pd.',
        'Informatika': 'Mersi, S.T.',
        'Seni Budaya': 'Rahmat Mas Said, S.Pd.',
        'B.Jawa': 'Fitri Lestari, S.S.',
    },
    '9B': {
        'IPA': 'Umi Kulstum, S.Pd.',
        'B.Indonesia': 'Maftuhah Rahayu, S.Pd.',
        'Matematika': 'Sri Purwanti, S.Pd.',
        'B.Inggris': 'Heni Purwaningsih, S.Pd.',
        'PJOK': 'Luthfan Qaedi W.',
        'PAI': 'Nurkhasanah, S.Ag.',
        'P.Pancasila': 'Feni Dwimartanti, S.Pd.',
        'IPS': 'Budi Prasetya, S.Pd.',
        'Informatika': 'Mersi, S.T.',
        'Seni Budaya': 'Rahmat Mas Said, S.Pd.',
        'B.Jawa': 'Fitri Lestari, S.S.',
    },
    '9C': {
        'IPA': 'Umi Kulstum, S.Pd.',
        'B.Indonesia': 'Maftuhah Rahayu, S.Pd.',
        'Matematika': 'Sri Purwanti, S.Pd.',
        'B.Inggris': 'Heni Purwaningsih, S.Pd.',
        'PJOK': 'Luthfan Qaedi W.',
        'PAI': 'Nurkhasanah, S.Ag.',
        'P.Pancasila': 'Feni Dwimartanti, S.Pd.',
        'IPS': 'Budi Prasetya, S.Pd.',
        'Informatika': 'Mersi, S.T.',
        'Seni Budaya': 'Rahmat Mas Said, S.Pd.',
        'B.Jawa': 'Fitri Lestari, S.S.',
    },
    '9D': {
        'IPA': 'Umi Kulstum, S.Pd.',
        'B.Indonesia': 'Maftuhah Rahayu, S.Pd.',
        'Matematika': 'Sri Purwanti, S.Pd.',
        'B.Inggris': 'Heni Purwaningsih, S.Pd.',
        'PJOK': 'Luthfan Qaedi W.',
        'PAI': 'Nurkhasanah, S.Ag.',
        'P.Pancasila': 'Feni Dwimartanti, S.Pd.',
        'IPS': 'Budi Prasetya, S.Pd.',
        'Informatika': 'Mersi, S.T.',
        'Seni Budaya': 'Rahmat Mas Said, S.Pd.',
        'B.Jawa': 'Fitri Lestari, S.S.',
    },
    '9E': {
        'IPA': 'Umi Kulstum, S.Pd.',
        'B.Indonesia': 'Maftuhah Rahayu, S.Pd.',
        'Matematika': 'Sri Purwanti, S.Pd.',
        'B.Inggris': 'Heni Purwaningsih, S.Pd.',
        'PJOK': 'Luthfan Qaedi W.',
        'PAI': 'Nurkhasanah, S.Ag.',
        'P.Pancasila': 'Feni Dwimartanti, S.Pd.',
        'IPS': 'Budi Prasetya, S.Pd.',
        'Informatika': 'Mersi, S.T.',
        'Seni Budaya': 'Rahmat Mas Said, S.Pd.',
        'B.Jawa': 'Fitri Lestari, S.S.',
    },
}

# ==========================================
# 3. MATRIKS MATA PELAJARAN MINGGUAN
# ==========================================
schedule_7 = {
    'Senin': {
        1: 'Matematika',
        2: 'Matematika',
        3: 'B.Indonesia',
        4: 'B.Indonesia',
        5: 'PAI',
        6: 'PAI',
        7: 'PAI',
        8: 'IPS',
        9: 'IPS',
    },
    'Selasa': {
        1: 'IPA',
        2: 'IPA',
        3: 'B.Inggris',
        4: 'B.Inggris',
        5: 'PJOK',
        6: 'PJOK',
        7: 'PJOK',
        8: 'B.Jawa',
        9: 'B.Jawa',
    },
    'Rabu': {
        1: 'Matematika',
        2: 'Matematika',
        3: 'B.Indonesia',
        4: 'B.Indonesia',
        5: 'IPA',
        6: 'IPA',
        7: 'Informatika',
        8: 'Informatika',
        9: 'Informatika',
    },
    'Kamis': {
        1: 'B.Inggris',
        2: 'B.Inggris',
        3: 'Matematika',
        4: 'IPA',
        5: 'B.Indonesia',
        6: 'B.Indonesia',
        7: 'P.Pancasila',
        8: 'P.Pancasila',
        9: 'Prakarya',
    },
    'Jumat': {
        1: 'Prakarya',
        2: 'Prakarya',
        3: 'P.Pancasila',
        4: 'IPS',
        5: 'IPS',
        6: 'BK',
    },
}

schedule_8 = {
    'Senin': {
        1: 'IPA',
        2: 'IPA',
        3: 'Matematika',
        4: 'Matematika',
        5: 'B.Indonesia',
        6: 'B.Indonesia',
        7: 'Seni Budaya',
        8: 'Seni Budaya',
        9: 'Seni Budaya',
    },
    'Selasa': {
        1: 'PJOK',
        2: 'PJOK',
        3: 'PJOK',
        4: 'B.Inggris',
        5: 'B.Inggris',
        6: 'B.Indonesia',
        7: 'B.Indonesia',
        8: 'P.Pancasila',
        9: 'P.Pancasila',
    },
    'Rabu': {
        1: 'Matematika',
        2: 'IPA',
        3: 'IPA',
        4: 'PAI',
        5: 'PAI',
        6: 'PAI',
        7: 'IPS',
        8: 'IPS',
        9: 'B.Jawa',
    },
    'Kamis': {
        1: 'B.Indonesia',
        2: 'B.Indonesia',
        3: 'B.Inggris',
        4: 'B.Inggris',
        5: 'Matematika',
        6: 'Matematika',
        7: 'Informatika',
        8: 'Informatika',
        9: 'Informatika',
    },
    'Jumat': {
        1: 'PAI',
        2: 'IPS',
        3: 'IPS',
        4: 'P.Pancasila',
        5: 'B.Jawa',
        6: 'BK',
    },
}

schedule_9 = {
    'Senin': {
        1: 'B.Inggris',
        2: 'B.Inggris',
        3: 'IPA',
        4: 'IPA',
        5: 'Matematika',
        6: 'Matematika',
        7: 'IPS',
        8: 'IPS',
        9: 'B.Jawa',
    },
    'Selasa': {
        1: 'Matematika',
        2: 'Matematika',
        3: 'B.Indonesia',
        4: 'B.Indonesia',
        5: 'PAI',
        6: 'PAI',
        7: 'PAI',
        8: 'Informatika',
        9: 'Informatika',
    },
    'Rabu': {
        1: 'PJOK',
        2: 'PJOK',
        3: 'PJOK',
        4: 'B.Inggris',
        5: 'B.Inggris',
        6: 'IPA',
        7: 'Informatika',
        8: 'IPS',
        9: 'IPS',
    },
    'Kamis': {
        1: 'IPA',
        2: 'IPA',
        3: 'Matematika',
        4: 'B.Indonesia',
        5: 'B.Indonesia',
        6: 'B.Indonesia',
        7: 'Seni Budaya',
        8: 'Seni Budaya',
        9: 'Seni Budaya',
    },
    'Jumat': {
        1: 'B.Indonesia',
        2: 'P.Pancasila',
        3: 'P.Pancasila',
        4: 'P.Pancasila',
        5: 'B.Jawa',
        6: 'BK',
    },
}

class_schedules = {}
for c in ['7A', '7B', '7C', '7D', '7E']:
  class_schedules[c] = schedule_7
for c in ['8A', '8B', '8C', '8D', '8E']:
  class_schedules[c] = schedule_8
for c in ['9A', '9B', '9C', '9D', '9E']:
  class_schedules[c] = schedule_9

classes = list(teaching_assignment.keys())

# ==========================================
# 4. DEKORASI & GAYA STYLE EXCEL
# ==========================================
navy_fill = PatternFill(
    start_color='1F497D', end_color='1F497D', fill_type='solid'
)
sub_header_fill = PatternFill(
    start_color='DCE6F1', end_color='DCE6F1', fill_type='solid'
)
time_fill = PatternFill(
    start_color='F2F2F2', end_color='F2F2F2', fill_type='solid'
)
break_fill = PatternFill(
    start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'
)

font_title = Font(name='Arial', size=14, bold=True, color='1F497D')
font_header = Font(name='Arial', size=10, bold=True, color='FFFFFF')
font_sub_header = Font(name='Arial', size=10, bold=True, color='1F497D')
font_bold = Font(name='Arial', size=9, bold=True)
font_regular = Font(name='Arial', size=9)
font_italic = Font(name='Arial', size=8, italic=True, color='595959')

thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)

align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
align_left = Alignment(horizontal='left', vertical='center')

# Slot Waktu Pelajaran
time_slots = [
    (
        'Senin',
        [
            (0, '07:00 - 07:20', 'Upacara'),
            (1, '07:20 - 08:00', 'Pembelajaran'),
            (2, '08:00 - 08:40', 'Pembelajaran'),
            (3, '08:40 - 09:20', 'Pembelajaran'),
            ('IST1', '09:20 - 09:40', 'Istirahat I'),
            (4, '09:40 - 10:20', 'Pembelajaran'),
            (5, '10:20 - 11:00', 'Pembelajaran'),
            (6, '11:00 - 11:40', 'Pembelajaran'),
            ('IST2', '11:40 - 12:20', 'Istirahat II'),
            (7, '12:20 - 13:00', 'Pembelajaran'),
            (8, '13:00 - 13:40', 'Pembelajaran'),
            (9, '13:40 - 14:20', 'Pembelajaran'),
        ],
    ),
    (
        'Selasa',
        [
            (0, '07:00 - 07:20', 'Pembiasaan'),
            (1, '07:20 - 08:00', 'Pembelajaran'),
            (2, '08:00 - 08:40', 'Pembelajaran'),
            (3, '08:40 - 09:20', 'Pembelajaran'),
            ('IST1', '09:20 - 09:40', 'Istirahat I'),
            (4, '09:40 - 10:20', 'Pembelajaran'),
            (5, '10:20 - 11:00', 'Pembelajaran'),
            (6, '11:00 - 11:40', 'Pembelajaran'),
            ('IST2', '11:40 - 12:20', 'Istirahat II'),
            (7, '12:20 - 13:00', 'Pembelajaran'),
            (8, '13:00 - 13:40', 'Pembelajaran'),
            (9, '13:40 - 14:20', 'Pembelajaran'),
        ],
    ),
    (
        'Rabu',
        [
            (0, '07:00 - 07:20', 'Pembiasaan'),
            (1, '07:20 - 08:00', 'Pembelajaran'),
            (2, '08:00 - 08:40', 'Pembelajaran'),
            (3, '08:40 - 09:20', 'Pembelajaran'),
            ('IST1', '09:20 - 09:40', 'Istirahat I'),
            (4, '09:40 - 10:20', 'Pembelajaran'),
            (5, '10:20 - 11:00', 'Pembelajaran'),
            (6, '11:00 - 11:40', 'Pembelajaran'),
            ('IST2', '11:40 - 12:20', 'Istirahat II'),
            (7, '12:20 - 13:00', 'Pembelajaran'),
            (8, '13:00 - 13:40', 'Pembelajaran'),
            (9, '13:40 - 14:20', 'Pembelajaran'),
        ],
    ),
    (
        'Kamis',
        [
            (0, '07:00 - 07:20', 'Pembiasaan'),
            (1, '07:20 - 08:00', 'Pembelajaran'),
            (2, '08:00 - 08:40', 'Pembelajaran'),
            (3, '08:40 - 09:20', 'Pembelajaran'),
            ('IST1', '09:20 - 09:40', 'Istirahat I'),
            (4, '09:40 - 10:20', 'Pembelajaran'),
            (5, '10:20 - 11:00', 'Pembelajaran'),
            (6, '11:00 - 11:40', 'Pembelajaran'),
            ('IST2', '11:40 - 12:20', 'Istirahat II'),
            (7, '12:20 - 13:00', 'Pembelajaran'),
            (8, '13:00 - 13:40', 'Pembelajaran'),
            (9, '13:40 - 14:20', 'Pembelajaran'),
        ],
    ),
    (
        'Jumat',
        [
            (0, '07:00 - 07:40', 'Pembiasaan'),
            (1, '07:40 - 08:10', 'Pembelajaran'),
            (2, '08:10 - 08:40', 'Pembelajaran'),
            (3, '08:40 - 09:10', 'Pembelajaran'),
            ('IST1', '09:10 - 09:30', 'Istirahat I'),
            (4, '09:30 - 10:00', 'Pembelajaran'),
            (5, '10:00 - 10:30', 'Pembelajaran'),
            (6, '10:30 - 11:00', 'Pembelajaran'),
        ],
    ),
]

# ==========================================
# 5. BUAT SHEET MASTER (Jadwal_Semua_Kelas)
# ==========================================
ws_master = wb.create_sheet(title='Jadwal_Semua_Kelas')
ws_master.merge_cells('A1:Q1')
ws_master['A1'] = 'JADWAL PELAJARAN SMP - SEMESTER GENAP'
ws_master['A1'].font = font_title

ws_master.merge_cells('A4:A5')
ws_master['A4'] = 'Hari'
ws_master.merge_cells('B4:B5')
ws_master['B4'] = 'Jam'
ws_master.merge_cells('C4:C5')
ws_master['C4'] = 'Waktu'

for col_idx in range(1, 4):
  c = ws_master.cell(row=4, column=col_idx)
  c.fill, c.font, c.alignment = navy_fill, font_header, align_center

ws_master.merge_cells('D4:Q4')
ws_master['D4'] = 'Rombongan Belajar (Kelas)'
ws_master['D4'].fill, ws_master['D4'].font, ws_master['D4'].alignment = (
    navy_fill,
    font_header,
    align_center,
)

for i, cls in enumerate(classes):
  c = ws_master.cell(row=5, column=4 + i, value=cls)
  c.fill, c.font, c.alignment = sub_header_fill, font_sub_header, align_center

current_row = 6
for day, slots in time_slots:
  day_start = current_row
  for jam, waktu, jenis in slots:
    ws_master.cell(row=current_row, column=1, value=day).alignment = align_center
    ws_master.cell(
        row=current_row,
        column=2,
        value=str(jam) if isinstance(jam, int) else '',
    ).alignment = align_center
    ws_master.cell(row=current_row, column=3, value=waktu).alignment = (
        align_center
    )

    if isinstance(jam, str) or jam == 0:
      for col_idx in range(1, 18):
        c = ws_master.cell(row=current_row, column=col_idx)
        c.fill, c.font, c.border = break_fill, font_italic, thin_border
      ws_master.cell(row=current_row, column=4, value=jenis).alignment = (
          align_center
      )
    else:
      for col_idx in range(1, 4):
        c = ws_master.cell(row=current_row, column=col_idx)
        c.fill, c.border = time_fill, thin_border
      for i, cls in enumerate(classes):
        mapel = class_schedules[cls][day].get(jam, '')
        guru = teaching_assignment[cls].get(mapel, '')
        c = ws_master.cell(row=current_row, column=4 + i)
        c.value = (
            f'{mapel}\n({guru.split(",")[0]})' if mapel else '-'
        )
        c.font, c.alignment, c.border = font_regular, align_center, thin_border
    current_row += 1
  ws_master.merge_cells(
      start_row=day_start,
      start_column=1,
      end_row=current_row - 1,
      end_column=1,
  )

# ==========================================
# 6. BUAT SHEET INDIVIDUAL (Kelas_7A - 9E)
# ==========================================
for cls in classes:
  ws_cls = wb.create_sheet(title=f'Kelas_{cls}')
  ws_cls.merge_cells('A1:E1')
  ws_cls['A1'] = f'JADWAL PELAJARAN KELAS {cls}'
  ws_cls['A1'].font = font_title

  headers = ['Hari', 'Jam Ke-', 'Waktu', 'Mata Pelajaran', 'Guru Pengampu']
  for col_idx, h in enumerate(headers, 1):
    c = ws_cls.cell(row=3, column=col_idx, value=h)
    c.fill, c.font, c.alignment = navy_fill, font_header, align_center

  row_c = 4
  for day, slots in time_slots:
    day_start = row_c
    for jam, waktu, jenis in slots:
      ws_cls.cell(row=row_c, column=1, value=day).alignment = align_center
      ws_cls.cell(
          row=row_c, column=2, value=str(jam) if isinstance(jam, int) else '-'
      ).alignment = align_center
      ws_cls.cell(row=row_c, column=3, value=waktu).alignment = align_center

      for col_i in range(1, 6):
        ws_cls.cell(row=row_c, column=col_i).border = thin_border

      if isinstance(jam, str) or jam == 0:
        ws_cls.merge_cells(
            start_row=row_c, start_column=4, end_row=row_c, end_column=5
        )
        c = ws_cls.cell(row=row_c, column=4, value=jenis)
        c.alignment, c.font = align_center, font_italic
        for col_i in range(1, 6):
          ws_cls.cell(row=row_c, column=col_i).fill = break_fill
      else:
        mapel = class_schedules[cls][day].get(jam, '-')
        guru = teaching_assignment[cls].get(mapel, '-')
        ws_cls.cell(row=row_c, column=4, value=mapel).font = font_bold
        ws_cls.cell(row=row_c, column=5, value=guru).font = font_regular
        for col_i in range(1, 4):
          ws_cls.cell(row=row_c, column=col_i).fill = time_fill
      row_c += 1
    ws_cls.merge_cells(
        start_row=day_start, start_column=1, end_row=row_c - 1, end_column=1
    )

# ==========================================
# 7. PENYESUAIAN LEBAR KOLOM & GRIDLINES
# ==========================================
for ws in wb.worksheets:
  ws.views.sheetView[0].showGridLines = True
  for col in ws.columns:
    max_len = max(
        len(line)
        for cell in col
        if cell.value
        for line in str(cell.value).split('\n')
    )
    col_letter = get_column_letter(col[0].column)
    ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

# ==========================================
# 8. SIMPAN FILE EXCEL
# ==========================================
output_filename = 'database_scheduler.xlsx'
wb.save(output_filename)
print(f'Selesai! Berhasil membuat file Excel jadwal: {output_filename}')
